from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from lxml import etree
from io import BytesIO
import pandas as pd
import zipfile
import os

def save_to_excel(dataframes, filename):
    """
    Saves a dictionary of pandas dataframes to an Excel file, with each
    dataframe on a separate sheet.
    
    Arguments:
    dataframes -- a dictionary where the keys are the sheet names and the values are pandas dataframes
    filename -- the name of the Excel file to save to
    """
    # Create a new Excel workbook and loop through the dataframes
    wb = Workbook()
    for sheet_name, df in dataframes.items():
        # Add a new sheet to the workbook and write the dataframe to it
        sheet = wb.create_sheet(title=sheet_name)
        for r, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c, value in enumerate(row):
                cell = sheet.cell(row=r+1, column=c+1)
                cell.value = value
                if r == 0:
                    cell.font = Font(bold=True)
        # Apply some formatting to the sheet
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        for row in sheet.rows:
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Save the workbook to disk
    wb.save(filename)

def get_table_elements(filename):
    """
    Reads an odt file and takes the tables inserted in them.
    
    Arguments:
    filename -- the name of the ODT file to read
    """
    with zipfile.ZipFile(filename) as odt:
        with odt.open('content.xml') as content:
            content_xml = content.read()

    #Create the xml object and get the root of the file
    xml_tree = etree.parse(BytesIO(content_xml))
    root = xml_tree.getroot()
    #Find all tag elements with table name
    table_elements = root.findall('.//table:table', namespaces=root.nsmap)
    return root, table_elements

def get_dataframes(root, table_elements):
    """
    Receive table elements from a xml file and return each table as pandas
    DataFrame.
    
    Arguments:
    root -- The root of xml file object
    table_elements -- The tables of XML file object
    """
    dataframes = []
    for table_element in table_elements:
        #First of all, get all rows of table element
        table_data = []
        row_elements = table_element.findall('.//table:table-row', namespaces=root.nsmap)
        #Now get all elements for each row
        for row_element in row_elements:
            row_data = []
            cell_elements = row_element.findall('.//table:table-cell', namespaces=root.nsmap)
            #Finally, get each cell element and append into a list
            for cell_element in cell_elements:
                text_content = cell_element.find('.//text:p', namespaces=root.nsmap)
                text_content = ''.join(text_content.itertext())
                row_data.append(text_content)
            table_data.append(row_data)
        #Save as pandas DataFrame
        df = pd.DataFrame(table_data, columns=table_data[0])
        df = df[1:]
        dataframes.append(df)
    return dataframes

#Getting only ODT files
filenames = [filename for filename in os.listdir() if filename.endswith(".odt")]

for i, filename in enumerate(filenames):
    root, table_elements = get_table_elements(filename)
    dataframes = get_dataframes(root, table_elements)

    dataframes = {f"Sheet_{i+1}": dataframe for i, dataframe in enumerate(dataframes)}
    save_to_excel(dataframes, f"planilhas/Excel_File_{i+1}.xlsx")